# -*- coding: utf-8 -*-
"""
Created on 2020.07
Updated on 2020.10
Author: MCHsu1
"""

"""
Part 1: Transfrom the ven_city & ven_cntry to ven_lat & ven_long
"""

from geopy.geocoders import Nominatim
from openpyxl import *
import pandas as pd
import pymssql
import time
import pandas.io.sql
import pymssql
import xlrd
import configparser

# read config.ini
def conf(a):
    config = configparser.ConfigParser()
    config.read(r'\\LABAP01\Emergency\02_Code\config.ini')
    config.sections()
    server = config['DEFAULT']['server']
    user = config['DEFAULT']['user']
    pwd = config['DEFAULT']['password']
    db = config['DEFAULT']['database']
    
    if a == 'server':
        return server
    elif a == 'user':
        return user
    elif a == 'pwd':
        return pwd
    elif a == 'db':
        return db

# Read data from excel and define a DataFrame
data = pd.read_excel (r'\\LABAP01\Emergency\User_upload\Vendors.xlsx')
df = pd.DataFrame(data, columns = ['venCity', 'venCntry'])

tStart = time.time()

# Load data, sheetname from excel
wd = load_workbook(r'\\LABAP01\Emergency\User_upload\Vendors.xlsx')
sheet = wd["Vendors"]

# Define a function for reture Lat, Long
def geo_lat_long(query): 
    geolocator = Nominatim(user_agent='lat_long_transform', timeout=10)
    loc = geolocator.geocode(query)
    lat = loc.latitude
    long = loc.longitude
    return lat, long

for i in range(0, df.shape[0]):
    CITY = df.iloc[i]['venCity']
    CNTRY = df.iloc[i]['venCntry']
 
    # Exception handling
    try:
        RESULT = str(CITY) + ' ' + str(CNTRY)
        print(i+1, RESULT)
        print(geo_lat_long(RESULT))
        print('')
    except AttributeError:
        print('')
        print('Exception: Error!' + '\n')
        continue
   
    #print(type(geo_lat_long(RESULT)))
    # Loop for the geo_lat_long(RESULT) and tranform into Strings
    Output = ','.join('%s' %id for id in geo_lat_long(RESULT))
   
    # Split a String into 2 strings
    LAT, LONG = Output.split(',', 1)
    #print("Lat:" + str(LAT) + '; ' + "Long:"+ str(LONG))
    #print('')    

    # Write the values of Lat, Long and save into Vendors.xlsx 
    wcell1       = sheet.cell(i+2, 5)
    wcell2       = sheet.cell(i+2, 6)
    wcell1.value = str(LAT)  # transform the data type into String
    wcell2.value = str(LONG) # transform the data type into String
    wd.save(r'\\LABAP01\Emergency\User_upload\Vendors.xlsx')
    
tEnd = time.time()
print ("It cost %f sec to complete the transformation of ven_lat & ven_long." % (tEnd - tStart)) # 會自動做進位(小數點後六位數)


"""
Part 2: Extract vendors data from excel to EDW STG
"""

# Create Connection and Cursor objects
conn = pymssql.connect(conf('server'), conf('user'), conf('pwd'), conf('db')) 
cursor = conn.cursor()  

# Read data from excel
data = pd.read_excel (r'\\LABAP01\Emergency\User_upload\Vendors.xlsx') 
#df = pd.DataFrame(data, columns= ['venCity', 'venCntry'])

# Open the workbook and define the worksheet
xlsx_data = xlrd.open_workbook(r'\\LABAP01\Emergency\User_upload\Vendors.xlsx')
sheet = xlsx_data.sheet_by_name("Vendors")

# Insert vendor's data into EDW
insert = '''
INSERT INTO WEC_EDW_DEV.STG.tbl_eas_vendor_tmp (
ven_id, ven_type, ven_name, ven_adr, ven_lat, ven_long, ven_city, ven_cntry, owner, ven_port, ven_contact, ven_code
) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) '''

# Delete dupicated rows
del_dup = '''
WITH cte AS (
SELECT *
FROM (select *, ROW_NUMBER() over (
partition by ven_id order by ven_id asc) row_num
from [WEC_EDW_DEV].[STG].[tbl_eas_vendor_tmp] ) a
where row_num not in ('1')
) DELETE FROM cte
'''

# Filter the blank value on venCity
filter_blank = '''
DELETE FROM [WEC_EDW_DEV].[STG].[tbl_eas_vendor_tmp]
WHERE ven_lat in ('') 
OR ven_long in ('') 
OR ven_city in ('')
'''

# grab existing row count in the database for validation later
cursor.execute("SELECT count(*) FROM [WEC_EDW_DEV].[STG].[tbl_eas_vendor_tmp]")
before_import = cursor.fetchone()


for r in range(1, sheet.nrows):
    venID       = sheet.cell(r,0).value
    venType     = sheet.cell(r,1).value
    venName     = sheet.cell(r,2).value
    venAddr     = sheet.cell(r,3).value
    venLat      = sheet.cell(r,4).value
    venLong     = sheet.cell(r,5).value
    venCity     = sheet.cell(r,6).value
    venCntry    = sheet.cell(r,7).value
    owner       = sheet.cell(r,8).value
    venAprt_Prt = sheet.cell(r,9).value
    venContact  = sheet.cell(r,10).value
    venCode     = sheet.cell(r,11).value
    
    values = (venID, venType, venName, venAddr, venLat , venLong, venCity, venCntry, owner, venAprt_Prt, venContact, venCode)
    
    cursor.execute(insert, values)
    cursor.execute(del_dup)
    cursor.execute(filter_blank)


# Commit the transaction
conn.commit()

# If we want to check if all rows are imported
# Successful notification    
cursor.execute("SELECT count(*) FROM [WEC_EDW_DEV].[STG].[tbl_eas_vendor_tmp]")
result = cursor.fetchone()

print((result[0] - before_import[0]) == len(data.index))  # should be True
print('Insert vendor data successfully!')
print('Delete duplicated rows!')

# Close the database connection
cursor.close()
conn.close()
